# -*- coding: utf-8 -*-
import filecmp
import os
import random
import urllib.request

import datetime
import urllib.error
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from mops_api.converter.converter import puwg92_do_wgs84
from mops_api.models import MOP, Operator

mops_api_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
NEW_XLSX_LOCATION = mops_api_dir + '/data/mopy_new.xlsx'
OLD_XLSX_LOCATION = mops_api_dir + '/data/mopy_old.xlsx'
LOCAL_CSV_LOCATION = ''


def type(x):
    x = str(x)
    if 'III' in x:
        return 3
    elif 'II' in x:
        return 2
    elif 'I' in x:
        return 1
    else:
        return 0


def randomize_free_places(mop):
    mop.taken_bus_dedicated_places = random.randint(0, mop.bus_dedicated_places)
    mop.taken_truck_places = random.randint(0, mop.truck_places)
    mop.taken_passenger_places = random.randint(0, mop.passenger_places)


def download_new_file(url):
    try:
        urllib.request.urlretrieve(url, NEW_XLSX_LOCATION)
    except urllib.error.HTTPError as e:
        print('Nie znaleziono na serwerze pliku o odpowiedniej nazwie')
        print(e)
        return False
    return True



def create_operator(attr):
    if not 'name' in attr.keys() or attr['permission'] == 'nie':
        return Operator(name='-')
    operator, _ = Operator.objects.get_or_create(name=attr['name'])
    for (key, value) in attr.items():
        if key != 'name' and value is not None:
            setattr(operator, key, value)
    operator.save()
    return operator


def create_mop(attr, _operator):
    if attr['title'] is not None and attr['x'] is not None and attr['y'] is not None:
        mop, created = MOP.objects.get_or_create(x=attr['x'], y=attr['y'], title=attr['title'], operator=_operator)
        for (key, value) in attr.items():
            if value is not None:
                setattr(mop, key, value)
        randomize_free_places(mop)
        mop.save()
        return mop


def phone_number(x):
    if x is None or x == '-':
        return '-'
    x = str(x)
    ret = ''
    counter = 0
    for c in x:
        if not c.isspace():
            counter += 1
            ret += c
        if counter == 9:
            break
    return ret


def email(x):
    if x is None or x == '-':
        return '-'
    return x.strip()


def name(x):
    if x is None or x == '-':
        return '-'
    return " ".join(x.split())


def boolean(x):
    if x is None or x == '-':
        return '-'
    return 'tak' in str(x)


mop_columns = {
    'number_in_excel': ('B', int),
    'department': ('C', str),
    'town': ('D', str),
    'title': ('E', str),
    'type': ('E', type),
    'x92': ('F', str),
    'y92': ('G', str),
    'road_technical_class': ('H', str),
    'road_number': ('I', str),
    'chainage': ('J', str),
    'direction': ('K', str),
    'turnoff': ('L', str),
    'passenger_places': ('M', int),
    'truck_places': ('N', int),
    'bus_dedicated_places': ('O', int),
    'security': ('P', boolean),
    'fence': ('Q', boolean),
    'monitoring': ('R', boolean),
    'lighting': ('S', boolean),
    'petrol_station': ('T', boolean),
    'dangerous_cargo_places': ('U', boolean),
    'restaurant': ('V', boolean),
    'sleeping_places': ('W', boolean),
    'toilets': ('X', boolean),
    'car_wash': ('Y', boolean),
    'garage': ('Z', boolean),
}

operator_columns = {
    'name': ('AA', str),
    'phone': ('AB', phone_number),
    'email': ('AC', email),
    'permission': ('AD', boolean)
}


def parse_field(prop, parser, row):
    try:
        return parser(prop)
    except Exception as e:
        print('Błąd parsowania w MOP-ie nr ' + str(row))
        print(e)


class Command(BaseCommand):
    @staticmethod
    def parse():

        wb = load_workbook(filename=NEW_XLSX_LOCATION, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=6):
            lp = row[1].value
            if lp is not None and isinstance(lp, int):
                mop_attr = {}
                operator_attr = {}

                for _key, (_col, _type) in mop_columns.items():
                    col = column_index_from_string(_col) - 1
                    mop_attr[_key] = parse_field(row[col].value, _type, lp)
                for _key, (_col, _type) in operator_columns.items():
                    col = column_index_from_string(_col) - 1
                    operator_attr[_key] = parse_field(row[col].value, _type, lp)

                mop_attr['x'], mop_attr['y'] = puwg92_do_wgs84(mop_attr['x92'], mop_attr['y92'])

                try:
                    operator = create_operator(operator_attr)
                    create_mop(mop_attr, operator)
                except Exception as e:
                    print('Błąd parsowania w MOP-ie nr ' + str(lp))
                    print(e)

        return True

    def handle(self, *args, **options):
        d = datetime.date.today()
        month_and_year = d.strftime('%m') + '.' + str(d.year)
        #XLSX_URL = 'https://www.gddkia.gov.pl/frontend/web/userfiles/articles/p/pliki-z-danymi-o-utrudnieniach_4395/MOP%20'
        #XLSX_URL += month_and_year + '%20final.xlsx'

        XLSX_URL = 'https://www.gddkia.gov.pl/frontend/web/userfiles/articles/p/pliki-z-danymi-o-utrudnieniach_4395/MOP%2011.2017%20final.xlsx'

        if download_new_file(XLSX_URL):
           # if filecmp.cmp(NEW_XLSX_LOCATION, OLD_XLSX_LOCATION):
            #    print('Plik identyczny z plikiem z poprzedniego miesiąca')
            #    os.remove(NEW_XLSX_LOCATION)
           # else:
                if self.parse():
                    os.rename(NEW_XLSX_LOCATION, OLD_XLSX_LOCATION)
                else:
                    print("Incorrect file on server")
